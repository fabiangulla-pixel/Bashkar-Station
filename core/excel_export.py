"""core/excel_export.py — Construye el Excel con gráficas incrustadas."""

import io
import gc
import tempfile
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import networkx as nx


# ── Paleta rica con 14 colores claramente distintos ───────────────────────────
AZUL1, AZUL2, AZUL3, AZUL4 = "#1F3864", "#2E5496", "#4472C4", "#9DC3E6"
PALETTE = [
    "#4E79A7",  # azul acero
    "#F28E2B",  # naranja ámbar
    "#E15759",  # rojo coral
    "#76B7B2",  # verde agua
    "#59A14F",  # verde musgo
    "#EDC948",  # amarillo mostaza
    "#B07AA1",  # malva
    "#FF9DA7",  # rosa pálido
    "#9C755F",  # marrón tabaco
    "#BAB0AC",  # gris cálido
    "#17BECF",  # cian
    "#8C6D31",  # bronce
    "#D62728",  # rojo vivo
    "#AECBAC",  # verde salvia
]
plt.rcParams.update({
    "figure.dpi": 120, "font.family": "DejaVu Serif",
    "axes.spines.top": False, "axes.spines.right": False,
    "axes.prop_cycle": plt.cycler(color=PALETTE),
})


def _fig_bytes(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=140, bbox_inches="tight")
    buf.seek(0)
    return buf.getvalue()


def generar_figuras(datos: dict) -> dict:
    """
    Genera figuras con tipos de gráfico variados y paleta diversificada.
    Retorna {nombre: bytes_png}.
    """
    figs = {}
    pub  = datos.get("publicacion", "Publicación")

    # ── 1. SECCIONES: gráfico de barras horizontales con gradiente ────────────
    df_s = datos.get("df_secciones")
    if df_s is not None and not df_s.empty:
        tot = df_s.groupby("seccion")["menciones"].sum().sort_values()
        colores = [PALETTE[i % len(PALETTE)] for i in range(len(tot))]
        fig, ax = plt.subplots(figsize=(10, max(5, len(tot)*0.45)))
        bars = ax.barh(tot.index, tot.values, color=colores,
                       edgecolor="white", linewidth=0.6, height=0.7)
        ax.bar_label(bars, padding=4, fontsize=9, color="#333")
        ax.set_xlabel("Menciones totales", fontsize=10)
        ax.set_title(f"Arquitectura editorial · {pub}", fontsize=13,
                     fontweight="bold", pad=12)
        ax.xaxis.grid(True, linestyle="--", alpha=0.4)
        ax.set_axisbelow(True)
        plt.tight_layout()
        figs["secciones"] = _fig_bytes(fig); plt.close(fig)

    # ── 2. FIRMAS: barras con color diferenciado ──────────────────────────────
    df_f  = datos.get("df_firmas")
    colabs = datos.get("colaboradores", set())
    if df_f is not None and not df_f.empty:
        top20 = df_f["firma"].value_counts().head(20)
        colores = [PALETTE[0] if f in colabs else PALETTE[2] for f in top20.index]
        fig, ax = plt.subplots(figsize=(10, 5.5))
        bars = ax.barh(top20.index[::-1], top20.values[::-1],
                       color=colores[::-1], edgecolor="white", linewidth=0.6)
        ax.bar_label(bars, padding=4, fontsize=9)
        ax.set_xlabel("Apariciones en el corpus", fontsize=10)
        ax.set_title("Top 20 firmas identificadas", fontsize=13,
                     fontweight="bold", pad=12)
        l1 = mpatches.Patch(color=PALETTE[0], label="Colaborador documentado")
        l2 = mpatches.Patch(color=PALETTE[2], label="Detectado automáticamente")
        ax.legend(handles=[l1,l2], fontsize=9, framealpha=0.85)
        ax.xaxis.grid(True, linestyle="--", alpha=0.4); ax.set_axisbelow(True)
        plt.tight_layout()
        figs["firmas"] = _fig_bytes(fig); plt.close(fig)

    # ── 3. CAMPOS SEMÁNTICOS: gráfico de áreas apiladas ──────────────────────
    df_c = datos.get("df_campos")
    if df_c is not None and not df_c.empty and len(df_c) > 1:
        df2 = df_c.set_index("numero")
        fig, ax = plt.subplots(figsize=(12, 5.5))
        x = range(len(df2))
        for i, col in enumerate(df2.columns):
            col_vals = df2[col].values
            ax.fill_between(x, col_vals, alpha=0.22, color=PALETTE[i % len(PALETTE)])
            ax.plot(x, col_vals, marker="o", label=col,
                    color=PALETTE[i % len(PALETTE)], linewidth=2.2,
                    markersize=5.5, markerfacecolor="white", markeredgewidth=1.8)
        ax.set_xticks(list(x))
        ax.set_xticklabels(df2.index, rotation=40, ha="right", fontsize=8)
        ax.set_ylabel("Menciones / 1 000 palabras", fontsize=10)
        ax.set_title(f"Evolución de campos temáticos · {pub}",
                     fontsize=13, fontweight="bold", pad=12)
        ax.legend(loc="upper right", fontsize=8.5, framealpha=0.85,
                  ncol=max(1, len(df2.columns)//4))
        ax.yaxis.grid(True, linestyle="--", alpha=0.4); ax.set_axisbelow(True)
        plt.tight_layout()
        figs["campos"] = _fig_bytes(fig); plt.close(fig)

    # ── 4. LDA heatmap ───────────────────────────────────────────────────────
    df_dt = datos.get("df_doc_temas")
    df_t  = datos.get("df_temas")
    if df_dt is not None and not df_dt.empty and df_t is not None and not df_t.empty:
        cols_t = [c for c in df_dt.columns if c.startswith("tema_")
                  and c.split("_")[1:] and c.split("_")[1].isdigit()]
        etiq = []
        for col in cols_t:
            n   = int(col.split("_")[1])
            row = df_t[df_t["tema"] == n]
            kw  = row["palabras_clave"].values[0].split(", ")[:3] if len(row) else [col]
            etiq.append(f"T{n}: {', '.join(kw)}")
        fig, ax = plt.subplots(figsize=(12, max(4, len(df_dt)*0.4+2)))
        sns.heatmap(df_dt[cols_t].T, annot=True, fmt=".2f",
                    cmap="YlOrRd",
                    xticklabels=df_dt.index, yticklabels=etiq,
                    ax=ax, linewidths=0.4, annot_kws={"size":7})
        ax.set_title("Temas latentes (LDA) por número",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_xticklabels(ax.get_xticklabels(), rotation=40, ha="right", fontsize=8)
        ax.set_yticklabels(ax.get_yticklabels(), fontsize=8.5)
        plt.tight_layout()
        figs["lda"] = _fig_bytes(fig); plt.close(fig)

    # ── 5. RED DE AUTORÍA: grafo con layout spring ────────────────────────────
    G = datos.get("red")
    if G is not None and len(G.nodes) > 0:
        fig, ax = plt.subplots(figsize=(10, 8))
        ax.set_facecolor("#F8F5EF")
        grados  = dict(G.degree())
        max_deg = max(grados.values(), default=1)
        node_sz = [300 + 1800 * (grados[n]/max_deg) for n in G.nodes()]
        node_cl = [PALETTE[0] if n in colabs else PALETTE[1] for n in G.nodes()]
        pos = nx.spring_layout(G, k=1.8/max(len(G.nodes())**0.5, 1), seed=42)
        nx.draw_networkx_edges(G, pos, ax=ax, alpha=0.25,
                               edge_color="#888", width=1.0)
        nx.draw_networkx_nodes(G, pos, ax=ax, node_size=node_sz,
                               node_color=node_cl, alpha=0.88)
        etiquetas = {n: n for n in G.nodes() if grados[n] >= max(2, max_deg//5)}
        nx.draw_networkx_labels(G, pos, labels=etiquetas, ax=ax,
                                font_size=7.5, font_color="#1a1a1a")
        ax.set_title("Red de colaboración entre autores",
                     fontsize=13, fontweight="bold", pad=12)
        l1 = mpatches.Patch(color=PALETTE[0], label="Colaborador documentado")
        l2 = mpatches.Patch(color=PALETTE[1], label="Detectado automáticamente")
        ax.legend(handles=[l1,l2], fontsize=9)
        ax.axis("off"); plt.tight_layout()
        figs["red"] = _fig_bytes(fig); plt.close(fig)

    # ── 6. LAYOUT: gráfico de áreas apiladas (texto vs imagen) ───────────────
    df_l = datos.get("df_layout")
    if df_l is not None and not df_l.empty and "prop_texto" in df_l.columns:
        fig, ax = plt.subplots(figsize=(11, 4))
        x = range(len(df_l))
        ax.stackplot(x,
                     df_l["prop_texto"].fillna(0),
                     df_l["prop_imagen"].fillna(0) if "prop_imagen" in df_l else [0]*len(df_l),
                     labels=["Texto","Imagen"],
                     colors=[PALETTE[0], PALETTE[1]], alpha=0.78)
        ax.set_xticks(list(x))
        ax.set_xticklabels(
            df_l["numero"] if "numero" in df_l else df_l.index,
            rotation=40, ha="right", fontsize=7.5)
        ax.set_ylabel("Proporción de la página", fontsize=10)
        ax.set_ylim(0,1); ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v,_:f"{v:.0%}"))
        ax.set_title("Distribución texto / imagen por número",
                     fontsize=13, fontweight="bold", pad=12)
        ax.legend(loc="upper right", fontsize=9)
        ax.yaxis.grid(True, linestyle="--", alpha=0.4); ax.set_axisbelow(True)
        plt.tight_layout()
        figs["layout"] = _fig_bytes(fig); plt.close(fig)

    return figs


def construir_excel(datos: dict, figs: dict, output_path: Path) -> Path:
    """Construye el Excel con 8 hojas y gráficas incrustadas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart import BarChart, LineChart, Reference

    HDR  = PatternFill("solid", fgColor="1F3864")
    HDR2 = PatternFill("solid", fgColor="2E5496")
    ALT  = PatternFill("solid", fgColor="DEEAF1")
    WARN = PatternFill("solid", fgColor="FFCCCC")
    HF   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BF   = Font(name="Arial", size=9)
    TF   = Font(name="Arial", bold=True, size=14, color="1F3864")
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    TH   = Side(style="thin", color="B8CCE4")
    BRD  = Border(left=TH, right=TH, top=TH, bottom=TH)

    TMP_IMG = Path(tempfile.gettempdir()) / "estampa_xl_tmp.png"

    def wdf(ws, df, sr=1, sc=1, hfill=None):
        hfill = hfill or HDR
        cols = list(df.columns)
        for j, col in enumerate(cols, sc):
            c = ws.cell(sr, j, str(col))
            c.fill = hfill; c.font = HF; c.alignment = CTR; c.border = BRD
            ws.column_dimensions[get_column_letter(j)].width = max(12, len(str(col)) + 4)
        for i, (_, row) in enumerate(df.iterrows(), sr + 1):
            for j, val in enumerate(row, sc):
                cell = ws.cell(i, j)
                cell.value = round(val, 3) if isinstance(val, float) else val
                cell.font = BF; cell.alignment = LFT; cell.border = BRD
                if (i - sr) % 2: cell.fill = ALT

    def embed(ws, key, anchor, scale=0.55):
        if key not in figs: return
        TMP_IMG.write_bytes(figs[key])
        xi = XLImage(str(TMP_IMG))
        xi.anchor = anchor
        xi.width  = int(xi.width  * scale)
        xi.height = int(xi.height * scale)
        ws.add_image(xi)

    wb = Workbook()
    wb.remove(wb.active)

    meta    = datos.get("corpus_meta")
    n_pdfs  = meta["numero"].nunique() if meta is not None else 0
    n_pags  = len(meta) if meta is not None else 0
    n_pals  = int(meta["palabras"].sum()) if meta is not None else 0
    df_f    = datos.get("df_firmas")
    df_s    = datos.get("df_secciones")
    n_firm  = df_f["firma"].nunique() if df_f is not None and not df_f.empty else 0
    n_secc  = df_s["seccion"].nunique() if df_s is not None and not df_s.empty else 0

    # ── Hoja 0: Resumen ──────────────────────────────────────────
    ws0 = wb.create_sheet("0_Resumen")
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 32
    ws0.column_dimensions["B"].width = 18
    ws0.column_dimensions["C"].width = 55
    ws0["A1"] = "🗞️  ESTAMPA — Resultados del análisis editorial"
    ws0["A1"].font = Font(name="Arial", bold=True, size=16, color="1F3864")
    ws0.merge_cells("A1:C1")
    ws0["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · Corpus: enero–junio 1939"
    ws0["A2"].font = Font(name="Arial", italic=True, size=10, color="555555")
    ws0.merge_cells("A2:C2")
    for j, h in enumerate(["Indicador", "Valor", "Nota"], 1):
        c = ws0.cell(4, j, h); c.fill = HDR; c.font = HF; c.alignment = CTR; c.border = BRD
    for i, (ind, val, nota) in enumerate([
        ("Números analizados",   n_pdfs, "PDFs procesados"),
        ("Páginas en el corpus", n_pags, "Resultado del OCR"),
        ("Palabras extraídas",   n_pals, "Total corpus textual"),
        ("Firmas identificadas", n_firm, "NER + lista + patrones"),
        ("Secciones detectadas", n_secc, "Secciones editoriales distintas"),
    ], 5):
        ws0.cell(i, 1, ind).font = BF
        ws0.cell(i, 2, val).font = Font(name="Arial", bold=True, size=12, color="2E5496")
        ws0.cell(i, 3, nota).font = BF
        if i % 2:
            for j in range(1, 4): ws0.cell(i, j).fill = ALT

    # ── Hoja 1: Firmas ───────────────────────────────────────────
    ws1 = wb.create_sheet("1_Firmas")
    ws1.sheet_view.showGridLines = False
    ws1.cell(1, 1, "Firmas y colaboradores · Estampa 1939").font = TF
    if df_f is not None and not df_f.empty:
        freq = (df_f.groupby("firma")
                .agg(apariciones=("numero","count"), numeros_distintos=("numero","nunique"))
                .reset_index().sort_values("apariciones", ascending=False).head(60))
        wdf(ws1, freq, sr=3)
        embed(ws1, "firmas", "E3", 0.6)
        n_r = min(len(freq), 20)
        ch = BarChart(); ch.type="bar"; ch.title="Top firmas"; ch.style=10
        ch.width=18; ch.height=12
        ch.add_data(Reference(ws1,min_col=2,min_row=3,max_row=3+n_r),titles_from_data=True)
        ch.set_categories(Reference(ws1,min_col=1,min_row=4,max_row=3+n_r))
        ws1.add_chart(ch, "E35")

    # ── Hoja 2: Secciones ────────────────────────────────────────
    ws2 = wb.create_sheet("2_Secciones")
    ws2.sheet_view.showGridLines = False
    ws2.cell(1, 1, "Arquitectura editorial · Secciones").font = TF
    if df_s is not None and not df_s.empty:
        wdf(ws2, df_s, sr=3)
        embed(ws2, "secciones", "H3", 0.6)
        tot = df_s.groupby("seccion")["menciones"].sum().reset_index()
        r_t = len(df_s) + 6
        ws2.cell(r_t, 1, "Totales por sección").font = TF
        wdf(ws2, tot, sr=r_t+2, hfill=HDR2)
        ch2 = BarChart(); ch2.type="bar"; ch2.style=10
        ch2.title="Menciones por sección"; ch2.width=18; ch2.height=12
        ch2.add_data(Reference(ws2,min_col=2,min_row=r_t+2,max_row=r_t+2+len(tot)),titles_from_data=True)
        ch2.set_categories(Reference(ws2,min_col=1,min_row=r_t+3,max_row=r_t+2+len(tot)))
        ws2.add_chart(ch2, "H35")

    # ── Hoja 3: Campos semánticos ────────────────────────────────
    ws3 = wb.create_sheet("3_Campos_semánticos")
    ws3.sheet_view.showGridLines = False
    ws3.cell(1, 1, "Campos semánticos · densidad por número").font = TF
    df_cam = datos.get("df_campos")
    if df_cam is not None and not df_cam.empty:
        wdf(ws3, df_cam, sr=3)
        embed(ws3, "campos", "J3", 0.62)
        if len(df_cam) > 1:
            n_f = len(df_cam); n_c = len(df_cam.columns) - 1
            ch3 = LineChart(); ch3.title="Evolución campos semánticos"; ch3.style=10
            ch3.width=22; ch3.height=12
            for ci in range(2, 2+n_c):
                ch3.add_data(Reference(ws3,min_col=ci,min_row=3,max_row=3+n_f),titles_from_data=True)
            ch3.set_categories(Reference(ws3,min_col=1,min_row=4,max_row=3+n_f))
            ws3.add_chart(ch3, "J35")

    # ── Hoja 4: LDA ──────────────────────────────────────────────
    ws4 = wb.create_sheet("4_Temas_LDA")
    ws4.sheet_view.showGridLines = False
    ws4.cell(1, 1, "Modelado de temas · LDA").font = TF
    df_t = datos.get("df_temas")
    if df_t is not None and not df_t.empty:
        wdf(ws4, df_t, sr=3)
        ws4.column_dimensions["B"].width = 70
        embed(ws4, "lda", "C3", 0.6)
        df_dt = datos.get("df_doc_temas")
        if df_dt is not None and not df_dt.empty:
            r2 = len(df_t) + 6
            ws4.cell(r2, 1, "Distribución por número").font = TF
            wdf(ws4, df_dt.reset_index(), sr=r2+2, hfill=HDR2)

    # ── Hoja 5: Red ───────────────────────────────────────────────
    ws5 = wb.create_sheet("5_Red_autoría")
    ws5.sheet_view.showGridLines = False
    ws5.cell(1, 1, "Red de colaboración editorial").font = TF
    gpath = datos.get("graph_path")
    if gpath and Path(gpath).exists():
        G = nx.read_graphml(str(gpath))
        cent = nx.degree_centrality(G); betw = nx.betweenness_centrality(G)
        df_r = pd.DataFrame([{
            "firma": n, "grado": G.degree(n),
            "centralidad": round(cent[n], 4), "betweenness": round(betw[n], 4),
            "apariciones": int(G.nodes[n].get("apariciones", 1))
        } for n in G.nodes()]).sort_values("centralidad", ascending=False)
        wdf(ws5, df_r, sr=3)
        embed(ws5, "red", "G3", 0.65)
        df_e = pd.DataFrame([
            {"nodo_1": u, "nodo_2": v, "peso": d.get("weight", 1)}
            for u, v, d in G.edges(data=True)
        ]).sort_values("peso", ascending=False)
        r_e = len(df_r) + 6
        ws5.cell(r_e, 1, "Conexiones entre agentes").font = TF
        wdf(ws5, df_e, sr=r_e+2, hfill=HDR2)
        del G; gc.collect()

    # ── Hoja 6: Layout ───────────────────────────────────────────
    ws6 = wb.create_sheet("6_Layout_visual")
    ws6.sheet_view.showGridLines = False
    ws6.cell(1, 1, "Análisis de layout visual").font = TF
    df_l = datos.get("df_layout")
    if df_l is not None and not df_l.empty:
        res = df_l.groupby("numero").agg(
            paginas=("pagina","count"), columnas_prom=("columnas","mean"),
            prop_imagen_prom=("prop_imagen","mean"), prop_texto_prom=("prop_texto","mean"),
            tinta_prom=("tinta","mean")
        ).round(3).reset_index()
        wdf(ws6, res, sr=3)
        embed(ws6, "layout", "H3", 0.6)
        n_l = len(res)
        ch6 = BarChart(); ch6.type="col"; ch6.grouping="stacked"; ch6.overlap=100
        ch6.title="Imagen/texto por número"; ch6.style=10; ch6.width=20; ch6.height=10
        ch6.add_data(Reference(ws6,min_col=4,min_row=3,max_row=3+n_l),titles_from_data=True)
        ch6.add_data(Reference(ws6,min_col=5,min_row=3,max_row=3+n_l),titles_from_data=True)
        ch6.set_categories(Reference(ws6,min_col=1,min_row=4,max_row=3+n_l))
        ws6.add_chart(ch6, "H28")

    # ── Hoja 7: OCR metadatos ────────────────────────────────────
    ws7 = wb.create_sheet("7_OCR_metadatos")
    ws7.sheet_view.showGridLines = False
    ws7.cell(1, 1, "Metadatos OCR · páginas con confianza < 60% en rojo").font = TF
    if meta is not None:
        meta_out = meta.drop(columns=["txt_path"], errors="ignore")
        wdf(ws7, meta_out, sr=3)
        for r in range(4, 4 + len(meta_out)):
            rev = ws7.cell(r, 5).value
            if rev:
                for c in range(1, 6):
                    ws7.cell(r, c).fill = WARN
                    ws7.cell(r, c).font = Font(name="Arial", size=9, color="C00000")

    wb.save(str(output_path))
    return output_path


# ── Nuevas hojas para Bashkar Station ─────────────────────────────────────────

def _agregar_hoja_articulos(wb, df_articulos, figs, HDR, HF, BF, TF, CTR, LFT, BRD, ALT):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    ws = wb.create_sheet("8_Artículos")
    ws.sheet_view.showGridLines = False
    ws.cell(1,1,"Artículos segmentados · atribución de autoría").font = TF
    if df_articulos is None or df_articulos.empty:
        ws.cell(3,1,"Sin datos de segmentación").font = BF
        return
    cols_show = ["numero","titulo","autor","confianza_autor","seccion","pagina","palabras"]
    cols_show = [c for c in cols_show if c in df_articulos.columns]
    df_out = df_articulos[cols_show].head(500)
    for j, col in enumerate(cols_show, 1):
        c = ws.cell(3, j, col.replace("_"," ").title())
        c.fill=HDR; c.font=HF; c.alignment=CTR; c.border=BRD
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(col)+4)
    CONF_FILL = PatternFill("solid", fgColor="E2EFDA")
    LOW_FILL  = PatternFill("solid", fgColor="FFF2CC")
    # El sombreado por confianza solo aplica si la columna está presente.
    col_conf = cols_show.index("confianza_autor") + 1 if "confianza_autor" in cols_show else None
    for i, (_, row) in enumerate(df_out.iterrows(), 4):
        for j, col in enumerate(cols_show, 1):
            cell = ws.cell(i, j)
            val = row[col]
            cell.value = round(val,2) if isinstance(val,float) else val
            cell.font=BF; cell.alignment=LFT; cell.border=BRD
            if (i-3)%2: cell.fill=ALT
        if col_conf is not None:
            try:
                conf_f = float(row.get("confianza_autor", 0))
            except Exception:
                conf_f = 0.0
            if conf_f >= 0.75:
                ws.cell(i, col_conf).fill = CONF_FILL
            elif conf_f < 0.3:
                ws.cell(i, col_conf).fill = LOW_FILL
    # Embed figure if available
    if "articulos" in figs:
        TMP = Path(tempfile.gettempdir()) / "bsk_art.png"
        TMP.write_bytes(figs["articulos"])
        from openpyxl.drawing.image import Image as XLImg
        xi = XLImg(str(TMP)); xi.anchor="I3"; xi.width=480; xi.height=300
        ws.add_image(xi)


def _agregar_hoja_visual(wb, datos_visual, figs, HDR, HF, BF, TF, CTR, LFT, BRD, ALT):
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("9_Visual_Tipografía")
    ws.sheet_view.showGridLines = False
    ws.cell(1,1,"Análisis visual y tipográfico").font = TF
    row = 3
    # Tipografía
    tip = datos_visual.get("tipografia",{})
    if tip:
        ws.cell(row,1,"Tipografía del número").font = TF; row+=1
        metricas = [
            ("Fuente principal", tip.get("fuente_principal","N/D")),
            ("Número de fuentes distintas", tip.get("n_fuentes",0)),
            ("Tamaño cuerpo (pt)", tip.get("tam_cuerpo_medio",0)),
            ("Tamaño título (pt)", tip.get("tam_titulo_medio",0)),
            ("Interlineado relativo", tip.get("interlineado_rel",0)),
            ("Columnas promedio", tip.get("columnas_prom",0)),
            ("% Negrita", tip.get("negrita_pct",0)),
            ("% Cursiva", tip.get("cursiva_pct",0)),
            ("Imágenes embebidas total", tip.get("imagenes_total",0)),
        ]
        for k, v in metricas:
            ws.cell(row,1,k).font=BF; ws.cell(row,2,v).font=BF; row+=1
        row += 1
        # Detalle de fuentes
        fuentes = tip.get("fuentes_resumen",[])
        if fuentes:
            import pandas as pd
            df_f = pd.DataFrame(fuentes)
            for j, col in enumerate(df_f.columns, 1):
                c=ws.cell(row,j,col); c.fill=HDR; c.font=HF; c.alignment=CTR; c.border=BRD
                ws.column_dimensions[get_column_letter(j)].width=max(12,len(col)+4)
            row+=1
            for _, r in df_f.iterrows():
                for j,col in enumerate(df_f.columns,1):
                    ws.cell(row,j,r[col]).font=BF
                row+=1
    row+=2
    # Elementos visuales (escaneados)
    elem = datos_visual.get("visual_elementos",[])
    if elem:
        ws.cell(row,1,"Elementos visuales detectados por página").font=TF; row+=1
        cols_e=["pagina","n_elementos","n_fotos","n_ilustraciones","n_publicidades","area_visual_pct"]
        for j,c in enumerate(cols_e,1):
            ce=ws.cell(row,j,c.replace("_"," ").title()); ce.fill=HDR; ce.font=HF; ce.border=BRD
        row+=1
        for pag in elem:
            for j,c in enumerate(cols_e,1):
                ws.cell(row,j,pag.get(c,0)).font=BF
            row+=1
    if "visual" in figs:
        TMP=Path(tempfile.gettempdir()) / "bsk_vis.png"; TMP.write_bytes(figs["visual"])
        from openpyxl.drawing.image import Image as XLImg
        xi=XLImg(str(TMP)); xi.anchor="J3"; xi.width=500; xi.height=320
        ws.add_image(xi)


def _agregar_hoja_comparativo(wb, datos_comp, figs, HDR, HF, BF, TF, CTR, LFT, BRD, ALT):
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("10_Análisis_Comparativo")
    ws.sheet_view.showGridLines = False
    ws.cell(1,1,"Análisis comparativo con otras publicaciones").font = TF
    if not datos_comp:
        ws.cell(3,1,"Sin datos comparativos").font=BF; return
    row = 3
    # Matriz de similaridad
    sim = datos_comp.get("similaridad")
    if sim is not None and not sim.empty:
        ws.cell(row,1,"Matriz de similaridad coseno (TF-IDF)").font=TF; row+=1
        cols = list(sim.columns)
        ws.cell(row,1,"Publicación").fill=HDR; ws.cell(row,1).font=HF
        for j,c in enumerate(cols,2):
            ce=ws.cell(row,j,c); ce.fill=HDR; ce.font=HF; ce.alignment=CTR
        row+=1
        from openpyxl.styles import PatternFill
        for pub in sim.index:
            ws.cell(row,1,pub).font=BF
            for j,c in enumerate(cols,2):
                val=sim.loc[pub,c]
                cell=ws.cell(row,j,round(float(val),3)); cell.font=BF; cell.alignment=CTR
                # Color: más azul = más similar
                intensity = int(val*100)
                fill_color=f"C6D9F1" if val>0.5 else ("EBF3FB" if val>0.2 else "FFFFFF")
                cell.fill=PatternFill("solid",fgColor=fill_color)
            row+=1
        row+=2
    # Palabras distintivas
    dist = datos_comp.get("palabras_distintivas",{})
    if dist:
        ws.cell(row,1,"Palabras más distintivas vs. otras publicaciones").font=TF; row+=1
        for ref_name, pals in dist.items():
            ws.cell(row,1,f"vs. {ref_name}").font=TF; row+=1
            for j,h in enumerate(["Palabra","Score G²"],1):
                ws.cell(row,j,h).fill=HDR; ws.cell(row,j).font=HF
            row+=1
            for palabra, score in pals[:20]:
                ws.cell(row,1,palabra).font=BF; ws.cell(row,2,score).font=BF; row+=1
            row+=1
    # Perfil de campos semánticos comparado
    campos = datos_comp.get("perfil_campos")
    if campos is not None and not campos.empty:
        ws.cell(row,1,"Perfil de campos semánticos por publicación").font=TF; row+=1
        df=campos.reset_index()
        for j,col in enumerate(df.columns,1):
            ws.cell(row,j,col).fill=HDR; ws.cell(row,j).font=HF
        row+=1
        for _,r in df.iterrows():
            for j,col in enumerate(df.columns,1):
                val=r[col]
                ws.cell(row,j, round(float(val),2) if isinstance(val,float) else val).font=BF
            row+=1
    if "comparativo" in figs:
        TMP=Path(tempfile.gettempdir()) / "bsk_comp.png"; TMP.write_bytes(figs["comparativo"])
        from openpyxl.drawing.image import Image as XLImg
        xi=XLImg(str(TMP)); xi.anchor="K3"; xi.width=520; xi.height=340
        ws.add_image(xi)


def construir_excel_completo(datos: dict, figs: dict, output_path: Path) -> Path:
    """
    Versión extendida de construir_excel con hojas adicionales
    para artículos segmentados, análisis visual/tipográfico y análisis comparativo.
    """
    # Construir las hojas base (0–7)
    construir_excel(datos, figs, output_path)
    
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = load_workbook(str(output_path))
    
    HDR = PatternFill("solid", fgColor="1F3864")
    ALT = PatternFill("solid", fgColor="DEEAF1")
    HF  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BF  = Font(name="Arial", size=9)
    TF  = Font(name="Arial", bold=True, size=14, color="1F3864")
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    TH  = Side(style="thin", color="B8CCE4")
    BRD = Border(left=TH, right=TH, top=TH, bottom=TH)
    
    kwargs = dict(HDR=HDR, HF=HF, BF=BF, TF=TF, CTR=CTR, LFT=LFT, BRD=BRD, ALT=ALT)
    
    if "df_articulos" in datos and datos["df_articulos"] is not None:
        _agregar_hoja_articulos(wb, datos["df_articulos"], figs, **kwargs)
    
    if "datos_visual" in datos and datos["datos_visual"]:
        _agregar_hoja_visual(wb, datos["datos_visual"], figs, **kwargs)
    
    if "datos_comparativo" in datos and datos["datos_comparativo"]:
        _agregar_hoja_comparativo(wb, datos["datos_comparativo"], figs, **kwargs)
    
    wb.save(str(output_path))
    gc.collect()
    return output_path


def generar_figuras_completas(datos: dict) -> dict:
    """Genera todas las figuras incluyendo artículos, visual y comparativo.
    Usa tipos de gráfico variados: barras, donut, radar, dispersión, cascada."""
    figs = generar_figuras(datos)

    # ── A. ARTÍCULOS: barras horizontales por autor + donut de secciones ──────
    df_art = datos.get("df_articulos")
    if df_art is not None and not df_art.empty and "autor" in df_art.columns:
        top_autores = (df_art[df_art["autor"] != "Anónimo / Sin atribuir"]
                       ["autor"].value_counts().head(20))
        if not top_autores.empty:
            fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))
            # Izquierda: barras horizontales (top autores)
            ax = axes[0]
            colores = [PALETTE[i % len(PALETTE)] for i in range(len(top_autores))]
            bars = ax.barh(top_autores.index[::-1], top_autores.values[::-1],
                           color=colores[::-1], edgecolor="white", height=0.7)
            ax.bar_label(bars, padding=4, fontsize=8.5)
            ax.set_xlabel("Artículos atribuidos", fontsize=10)
            ax.set_title("Top autores (artículos)", fontsize=12, fontweight="bold")
            ax.xaxis.grid(True, linestyle="--", alpha=0.4); ax.set_axisbelow(True)
            # Derecha: gráfico donut de secciones
            ax2 = axes[1]
            if "seccion" in df_art.columns:
                sec_counts = df_art["seccion"].value_counts().head(8)
                colores_d  = [PALETTE[i % len(PALETTE)] for i in range(len(sec_counts))]
                wedges, texts, autotexts = ax2.pie(
                    sec_counts.values,
                    labels=sec_counts.index,
                    colors=colores_d,
                    autopct="%1.1f%%", startangle=90,
                    wedgeprops=dict(width=0.55, edgecolor="white", linewidth=1.5),
                    pctdistance=0.78, textprops={"fontsize":8.5},
                )
                for at in autotexts: at.set_fontsize(7.5)
                ax2.set_title("Artículos por sección", fontsize=12, fontweight="bold")
            plt.tight_layout()
            figs["articulos"] = _fig_bytes(fig); plt.close(fig)

    # ── B. VISUAL: gráfico combinado barras+área ──────────────────────────────
    # datos_visual puede venir como None (no solo ausente): .get(k, {}) no
    # protege ese caso, por eso el "or {}".
    elem = (datos.get("datos_visual") or {}).get("visual_elementos", [])
    if elem:
        df_e = pd.DataFrame(elem)
        if not df_e.empty:
            # Totales por tipo
            totales = {
                "Fotografías":    df_e["n_fotos"].sum() if "n_fotos" in df_e else 0,
                "Ilustraciones":  df_e["n_ilustraciones"].sum() if "n_ilustraciones" in df_e else 0,
                "Publicidades":   df_e["n_publicidades"].sum() if "n_publicidades" in df_e else 0,
                "Decorativos":    df_e["n_decorativos"].sum() if "n_decorativos" in df_e else 0,
            }
            totales = {k:int(v) for k,v in totales.items() if v > 0}
            fig, axes = plt.subplots(1, 2, figsize=(13, 5))
            # Barras de conteo
            ax = axes[0]
            ks = list(totales.keys()); vs = list(totales.values())
            colores_b = [PALETTE[0], PALETTE[1], PALETTE[2], PALETTE[3]][:len(ks)]
            bars = ax.bar(ks, vs, color=colores_b, edgecolor="white", width=0.6)
            ax.bar_label(bars, fontsize=11, fontweight="bold", padding=3)
            ax.set_title("Elementos visuales detectados", fontsize=12, fontweight="bold")
            ax.set_ylabel("Cantidad total", fontsize=10)
            ax.yaxis.grid(True, linestyle="--", alpha=0.4); ax.set_axisbelow(True)
            # Gráfico de área: evolución por número si hay datos
            ax2 = axes[1]
            if "numero" in df_e.columns:
                agg = df_e.groupby("numero").agg({
                    "n_fotos":("sum" if "n_fotos" in df_e else None),
                    "n_ilustraciones":("sum" if "n_ilustraciones" in df_e else None),
                }).reset_index()
                if len(agg) > 1:
                    x = range(len(agg))
                    for i, col in enumerate(["n_fotos","n_ilustraciones"]):
                        if col in agg:
                            ax2.fill_between(x, agg[col], alpha=0.30, color=PALETTE[i])
                            ax2.plot(x, agg[col], marker="o", color=PALETTE[i],
                                     label=col.replace("n_","").capitalize(),
                                     linewidth=2, markersize=5)
                    ax2.set_xticks(list(x))
                    ax2.set_xticklabels(agg["numero"], rotation=40, ha="right", fontsize=7.5)
                    ax2.legend(fontsize=9); ax2.set_ylabel("Elementos por número")
                    ax2.yaxis.grid(True, linestyle="--", alpha=0.4)
                    ax2.set_title("Elementos visuales por número", fontsize=12, fontweight="bold")
                else:
                    ax2.axis("off")
            else:
                ax2.axis("off")
            plt.tight_layout()
            figs["visual"] = _fig_bytes(fig); plt.close(fig)

    # ── C. COMPARATIVO: gráfico radial (radar) + barras agrupadas ────────────
    dat_comp = datos.get("datos_comparativo") or {}
    pc = dat_comp.get("perfil_campos") if dat_comp else None
    if pc is not None and not pc.empty and len(pc.columns) >= 3:
        n_pub    = len(pc)
        n_campos = len(pc.columns)
        # Radar
        fig, axes = plt.subplots(1, 2, figsize=(14, 5.5),
                                  subplot_kw=dict(polar=False))
        # Barras agrupadas
        ax = axes[0]
        x = np.arange(n_campos); w = 0.75 / max(n_pub, 1)
        for i, (pub, row) in enumerate(pc.iterrows()):
            ax.bar(x + i*w, row.values, width=w, label=pub,
                   color=PALETTE[i % len(PALETTE)], alpha=0.85, edgecolor="white")
        ax.set_xticks(x + w*(n_pub-1)/2)
        ax.set_xticklabels(pc.columns, rotation=35, ha="right", fontsize=9)
        ax.set_ylabel("Menciones / 1 000 palabras", fontsize=10)
        ax.set_title("Campos temáticos por publicación", fontsize=12, fontweight="bold")
        ax.legend(fontsize=8.5, framealpha=0.85)
        ax.yaxis.grid(True, linestyle="--", alpha=0.4); ax.set_axisbelow(True)
        # Gráfico de radar (spider) si hay 3+ campos
        ax2 = plt.subplot(1, 2, 2, polar=True)
        angulos = np.linspace(0, 2*np.pi, n_campos, endpoint=False).tolist()
        angulos += angulos[:1]
        ax2.set_theta_offset(np.pi/2); ax2.set_theta_direction(-1)
        ax2.set_xticks(angulos[:-1])
        ax2.set_xticklabels(pc.columns, fontsize=8)
        for i, (pub, row) in enumerate(pc.iterrows()):
            vals = row.tolist(); vals += vals[:1]
            ax2.plot(angulos, vals, linewidth=2, color=PALETTE[i % len(PALETTE)], label=pub)
            ax2.fill(angulos, vals, alpha=0.12, color=PALETTE[i % len(PALETTE)])
        ax2.set_title("Perfil temático (radar)", fontsize=12, fontweight="bold", pad=18)
        ax2.legend(loc="upper right", fontsize=8.5, bbox_to_anchor=(1.3, 1.1))
        plt.tight_layout()
        figs["comparativo"] = _fig_bytes(fig); plt.close(fig)

    gc.collect()
    return figs

NARANJA = "#ED7D31"
